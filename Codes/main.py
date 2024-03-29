import os
import subprocess
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
import openpyxl as op


# global variables, command_path is for the binary exe command call address, input_size in the size of input data and key, scan_size is the size of scan chain
command_path = "./aes_scan_exam_sw5822_windows_amd64.exe"
input_size = 128
scan_size = 256

# input register bit position
def input_toExcel(data, fileName):  
    wb = op.Workbook()  # create object
    ws = wb.active
    ws.title = "scanchain_table"

    row = 0
    for key, value in data.items():
        row = key + 1
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value='input[' + str(value) + ']')
    

    wb.save(fileName)

#  data register bit position
def data_toExcel(data, num, fileName):  
    wb = op.load_workbook(fileName)  # create object
    ws = wb["scanchain_table"]

    row = 0
    for key, value in enumerate(data):
        row = value + 1
        ws.cell(row=row, column=1, value=value)
        ws.cell(row=row, column=2, value='data[' + str((num%4)*32 + key) + ']')
    
    wb.save(fileName)

# This function converts an array of bits to a bit string
def bit_array_to_bit_string(array):
    bit_string = ""
    for i in range(len(array)):
        bit_string += str(array[i])
    return bit_string

# This funciton converts a bit string to an array of bits
def bit_string_to_bit_array(string):
    bit_array=[]
    for i in range(len(string)):
        bit_array.append(int(string[i]))
    return bit_array

# This function converts an array of bits to an array of hex values
def bit_array_to_hex_array(array):
    size=int(len(array)/4)
    hex_array = [0 for i in range(size)]
    for i in range(size):
        for j in range(4):
            hex_array[i] += array[4 * i + j] << (3 - j)
        hex_array[i] = (hex(hex_array[i]))[2]
    return hex_array

# This function converts an array of hex values to a hex string
def hex_array_to_hex_string(array):
    hex_string = ""
    for i in range(32):
        hex_string += str(array[i])
    return hex_string

# This function extracts input register indices and input register bit map in the scan chain
def extract_input_indices_in_scan() :
    print("Start to analyze the input indices mapping to the scan chain bits...\n")
    print("Start to analyze the input indices mapping to the scan chain bits...\n", file=output_file)

    # specifying parameters for subprocess call of the exe file provided for the midterm to run the exe file and get the output for further processing
    command_input = "-input="
    command_clocks = "-clocks=2"
    command_emit = "-emit_scan"
    command_scan = "-scan_only"

    # the array of input indices
    input_indices = []
    input_indices_dic = {}
    input_indices_dic_sorted_by_scan = {}

    # for loop on all the input bits to change them and observe the change in the first clock cycle
    for k in range(input_size):
        input = [0 for i in range(input_size)]  # set input to all 0 first
        input[input_size - 1 - k] = 1           # set the LSB to 1 and gradually move to the left

        input_hex = bit_array_to_hex_array(input)
        #print("input plaintext " + str(k) + " round: ", input_hex)

        input_hex_str = hex_array_to_hex_string(input_hex)
        # print(input_hex_str)

        # the subprocess call executes the exe file and gets the output for more processing
        command_input = "-input=" + input_hex_str

        # return the output from the binary program
        res = subprocess.check_output([command_path, command_input, command_clocks, command_emit, command_scan])
        res = (str(res))[13:-3]
        # We just need the first clock cycle output of the scan chain for input indices mapping
        clk1_chain = res[:256]
        clk2_chain = res[269:]
        # print("output scan chain " + str(k) + " round: \n", res)


        # Finding which bit in the output is changed after changing one bit in the input in each round
        for i in range(scan_size):
            if(clk1_chain[i]=='1'):
                input_indices.append(i)
                input_indices_dic[k]=i

    # Generating the sorted indices based on the input register and the scan chain bits
    input_indices_sorted = input_indices.copy()
    input_indices_sorted.sort()

    for i in range(len(input_indices_sorted)):
        for j in range(len(input_indices_dic)):
            if(input_indices_dic[j]==input_indices_sorted[i]):
                input_indices_dic_sorted_by_scan[input_indices_sorted[i]]=j

    #print("Found the input indices in the scan chain ordered by input bits from LSB to MSB:\n",input_indices)
    #print("Found the input indices in the scan chain ordered by input bits from LSB to MSB:\n",input_indices,file=output_file)
    print("The input indices dictionary in the scan chain ordered by input bits from LSB to MSB(input reg loc:scan chain loc):\n", input_indices_dic)
    print("The input indices dictionary in the scan chain ordered by input bits from LSB to MSB(input reg loc:scan chain loc):\n", input_indices_dic,file=output_file)
    print("\nThe input indices dictionary in the scan chain ordered by scan bits from LSB to MSB(scan chain loc: input reg loc):\n",input_indices_dic_sorted_by_scan)
    print("\nThe input indices dictionary in the scan chain ordered by scan bits from LSB to MSB(scan chain loc: input reg loc):\n",input_indices_dic_sorted_by_scan, file=output_file)

    # print the input register bit positions to excel table
    input_toExcel(input_indices_dic_sorted_by_scan, 'scan_sheet.xlsx')
    
    return input_indices

# This function finds the RK0 and so the original key of the AES algorithm by doing the calculations
def extract_key():
    print("\nStart to analyze the byte candidates for RK0 and the main key of the AES algorithm...\n")
    print("\nStart to analyze the byte candidates for RK0 and the main key of the AES algorithm...\n", file=output_file)

    # Correspoding constants for the subprocess call of the exe file for running the algorithm
    command_input = "-input="
    command_clocks = "-clocks=2"
    command_emit = "-emit_scan"
    command_scan = "-scan_only"

    # Getting the initial_pattern at the second clock cycle by giving all zero pattern to the input
    input = [0 for i in range(input_size)]  # set input to all 0 first
    input_hex = bit_array_to_hex_array(input)
    input_hex_str = hex_array_to_hex_string(input_hex)
    command_input = "-input=" + input_hex_str
    res = subprocess.check_output([command_path, command_input, command_clocks, command_emit, command_scan])
    res = (str(res))[13:-3]
    clk1_chain = res[:256]
    clk2_chain = res[269:]
    # use the second clock cycle scan chain
    initial_pattern = bit_string_to_bit_array(clk2_chain)
    # RK0_byte_candidate is the array of candidate keys for each byte of the key which could have 2 different values
    RK0_byte_candidate=[[] for i in range(16)]

    # We have a loop on all 16 bytes of the input and key to do the mathematical calculations
    print("Start the loop on all bytes of input and key registers to find the key candidate bytes...\n")
    print("Start the loop on all bytes of input and key registers to find the key candidate bytes...\n", file=output_file)
    for b in range(16):
        # This is the array of data register indices in the scan chain for the corresponding byte which has 32 elements
        Data_reg_indices_in_scan = []

        print("Start to find data register indices in scan chain for byte number", b, "...")
        print("Start to find data register indices in scan chain for byte number", b, "...", file=output_file)
        # In this loop we check all the different combinations of the input data to produce a new pattern and find all the 32 bits of the corresponding data register indices in the scan output
        for pn in range(256):
            # Generate the input
            input = [0 for i in range(input_size)]
            for j in range(8):
                index = b * 8 + j
                input[input_size - 1 - index] = (pn >> j) & 0x01
                
            input_hex = bit_array_to_hex_array(input)
            input_hex_str = hex_array_to_hex_string(input_hex)
            command_input = "-input=" + input_hex_str
            res = subprocess.check_output([command_path, command_input, command_clocks, command_emit, command_scan])
            res = (str(res))[13:-3]
            clk1_chain = res[:256]
            clk2_chain = res[269:]
            new_pattern = bit_string_to_bit_array(clk2_chain)

            # Xor the initial and new patterns
            pattern_xor = [bit1 ^ bit2 for bit1, bit2 in zip(initial_pattern, new_pattern)]

            # Find the changed bits in the scan chain that are not in the input indices bits and add to the data register indices
            for i in range(scan_size):
                if (pattern_xor[i] == 1 and not (i in input_indices_in_scan) and not (i in Data_reg_indices_in_scan)):
                    Data_reg_indices_in_scan.append(i)
                   
        print("Data register indices in scan chain for input byte number", b, ":\n", Data_reg_indices_in_scan)
        print("Data register indices in scan chain for input byte number", b, ":\n", Data_reg_indices_in_scan, file=output_file)

        # print the data register bit positions to excel table
        data_toExcel(Data_reg_indices_in_scan, b, 'scan_sheet.xlsx')

        print("Start to find candidate key bytes for byte number", b, "...")
        print("Start to find candidate key bytes for byte number", b, "...", file=output_file)
        # This loop produces 2m and 2m+1 inputs and checks the scan chain corresponding data register indices to finally find the RK0 after XOR register
        for pn in range(128):
            # Generating inputs with 2m and 2m+1 patterns
            input_1 = [0 for i in range(input_size)]
            input_2 = [0 for i in range(input_size)]
            for j in range(8):
                index = b * 8 + j
                input_1[input_size - 1 - index] = ((2 * pn) >> j) & 0x01    # 2m
                input_2[input_size - 1 - index] = ((2 * pn + 1) >> j) & 0x01  # 2m + 1
          
            input_1_hex = bit_array_to_hex_array(input_1)
            input_1_hex_str = hex_array_to_hex_string(input_1_hex)
            input_2_hex = bit_array_to_hex_array(input_2)
            input_2_hex_str = hex_array_to_hex_string(input_2_hex)

            command_input = "-input=" + input_1_hex_str
            res = subprocess.check_output([command_path, command_input, command_clocks, command_emit, command_scan])
            res = (str(res))[13:-3]
            clk1_chain = res[:256]
            clk2_chain = res[269:]
            pattern_1 = bit_string_to_bit_array(clk2_chain)

            command_input = "-input=" + input_2_hex_str
            res = subprocess.check_output([command_path, command_input, command_clocks, command_emit, command_scan])
            res = (str(res))[13:-3]
            clk1_chain = res[:256]
            clk2_chain = res[269:]
            pattern_2 = bit_string_to_bit_array(clk2_chain)

            # Count number of ones in the xor of two patterns
            one_cnt = 0
            for i in range(32):
                pattern_xor_bit = pattern_1[Data_reg_indices_in_scan[i]] ^ pattern_2[Data_reg_indices_in_scan[i]]
                
                one_cnt += pattern_xor_bit
            # Check number of one values to guess the RK0 after XOR register according to the table
            if (one_cnt in (9,12,23,24)):
                if(one_cnt == 9):
                    value = 226
                elif(one_cnt == 12):
                    value = 242
                elif (one_cnt == 23):
                    value = 122
                elif (one_cnt == 24):
                    value = 130
                RK0_xor_1 = [0 for i in range(8)]
                RK0_xor_2 = [0 for i in range(8)]
                RK0_byte_1 = [0 for i in range(8)]
                RK0_byte_2 = [0 for i in range(8)]
                # Find the RK0 after XOR candidates
                for j in range(8):
                    RK0_xor_1[8 - 1 - j] = ((value) >> j) & 0x01
                    RK0_xor_2[8 - 1 - j] = ((value+1) >> j) & 0x01
                # Find the RK0 candidate bytes b(register value)^a(input)
                for j in range(8):
                    index = b * 8 + j
                    RK0_byte_1[8 - 1 - j] = RK0_xor_1[8 - 1 - j] ^ input_1[input_size - 1 - index]
                    RK0_byte_2[8 - 1 - j] = RK0_xor_2[8 - 1 - j] ^ input_1[input_size - 1 - index]
                    
                # Appending the RK0 candidate bytes to the RK0_byte_candidate list
                RK0_byte_candidate[16-1-b].append(RK0_byte_1)
                RK0_byte_candidate[16-1-b].append(RK0_byte_2)
                break
        print("Found candidate key bytes for byte number", b, "\n")
        print("Found candidate key bytes for byte number", b, "\n", file=output_file)

    print("Found each byte candidates for RK0 and the main key of the AES algorithm, RK0 byte candidates:\n", RK0_byte_candidate)
    print("Found each byte candidates for RK0 and the main key of the AES algorithm, RK0 byte candidates:\n", RK0_byte_candidate, file=output_file)
    return RK0_byte_candidate

# main function for AES scan chain attack
if __name__ == '__main__':

    # The path of output file for printing all the output results
    if(os.path.isfile("output.txt")):
        os.remove("output.txt")
    
    output_file = open('output.txt', 'w')
    print("Start the AES scan chain attack...\n")
    print("Start the AES scan chain attack...\n", file=output_file)

    # extract the input register indices in the scan chain 
    input_indices_in_scan = extract_input_indices_in_scan()

    # find the RK0 candidates for each byte of the key using the extract_key function
    RK0_byte_candidates = extract_key()

    # brute force search the final key
    print("\nStart the brute force search on the all key combinations...\n")
    print("\nStart the brute force search on the all key combinations...\n", file=output_file)

    # find out the ciphertext of plaintext 00000000000000000000000000000000 for brute force search
    plaintext = "00000000000000000000000000000000"
    command_in = "-input=" + plaintext
    ciphertxt = subprocess.check_output([command_path,command_in])
    ciphertxt = (str(ciphertxt))[9:-3]
    print("ciphertext: ", ciphertxt)
    print("ciphertext: ", ciphertxt, file=output_file)

    # test all the possible combinations of 16 bytes which have 2 candidates each so we have 2^16 different combinations
    for k in range(2**16):
        # set the value of the key_candidate using RK0_byte_candidates
        key_candidate = [0 for i in range(input_size)]
        for i in range(16):
            for j in range(8):
                index = i * 8 + j
                key_candidate[input_size - 1 - index] = RK0_byte_candidates[16-1-i][(k>>i)&0x01][8-1-j]

        key_candidate_hex_str = hex_array_to_hex_string(bit_array_to_hex_array(key_candidate))
        
        key_candidate_bytes = bytes.fromhex(key_candidate_hex_str)
        cipher = Cipher(algorithms.AES(key_candidate_bytes), modes.ECB())
        encryptor = cipher.encryptor()
        # Encrypt the plaintext using the candidae key
        ct = encryptor.update(bytes.fromhex(plaintext)) + encryptor.finalize()
        # print(str(ct.hex()))
        # At this point we compare the value of the ciphertext with the value of exe binary file output for all zero plaintext
        # By doing this brute force we find the final_key
        if(str(ct.hex())==ciphertxt.lower()):
            final_key = key_candidate
            print("Found the final Key: ", key_candidate_hex_str)
            print("Found the final Key: ", key_candidate_hex_str, file=output_file)

    output_file.close()